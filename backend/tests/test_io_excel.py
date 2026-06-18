import io
from datetime import date
import openpyxl
from app.io_parse import parse_deals_sheet


def _make_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "①案件日付別管理"
    ws.append(["売上月", "実施日", "代理店", "企業名", "研修名", "講師",
               "研修費用", "交通費", "その他", "消費税", "請求額", "講師料",
               "入金予定日", "サポートスタッフ"])
    ws.append([date(2026, 4, 30), date(2026, 4, 8), None, "サンライズ", None, "平野",
               450000, None, None, 45000, 495000, None, None, None])
    ws.append([date(2026, 7, 31), date(2026, 7, 30), None, "タダノ共栄会", "安全衛生", None,
               300000, 50000, None, 30000, 380000, 100000, None, None])
    return wb


def test_parse_falls_back_to_anken_nyuryoku_when_hizukebetsu_empty():
    # 「①案件日付別管理」が空で「案件入力」にデータがある年度ファイルに対応する
    wb = openpyxl.Workbook()
    empty = wb.active
    empty.title = "①案件日付別管理"
    empty.append(["売上月", "実施日", "代理店", "企業名", "研修名", "講師",
                  "研修費用", "交通費", "その他", "消費税", "請求額", "講師料",
                  "入金予定日", "サポートスタッフ"])
    src = wb.create_sheet("案件入力")
    src.append(["売上月", "実施日", "代理店", "企業名", "研修名", "講師",
                "研修費用", "交通費", "その他", "消費税", "請求額", "講師料",
                "入金予定日", "サポートスタッフ"])
    src.append([date(2025, 5, 31), date(2025, 5, 23), None, "中小企業大学校", "レジリエンス",
                "高橋", 36000, None, None, 3600, 39600, None, None, None])
    deals = parse_deals_sheet(wb)
    assert len(deals) == 1
    assert deals[0].client == "中小企業大学校"
    assert deals[0].fee == 36000
    assert deals[0].held_on == date(2025, 5, 23)


def test_parse_deals_sheet_maps_rows():
    wb = _make_workbook()
    deals = parse_deals_sheet(wb)
    assert len(deals) == 2
    d0 = deals[0]
    assert d0.client == "サンライズ"
    assert d0.fee == 450000
    assert d0.tax == 45000
    assert d0.billing == 495000
    assert d0.held_on == date(2026, 4, 8)
    d1 = deals[1]
    assert d1.transport == 50000
    assert d1.instructor_fee == 100000


def test_import_endpoint_loads_and_wipes(client):
    wb = _make_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    res = client.post(
        "/api/import/excel?wipe=true",
        files={"file": ("test.xlsx", buf,
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    assert res.json()["imported"] == 2
    deals = client.get("/api/deals", params={"fiscal_year": 2026}).json()
    assert len(deals) == 2
    clients = client.get("/api/masters/clients").json()
    assert any(m["name"] == "サンライズ" for m in clients)


def test_export_endpoint_returns_xlsx(client):
    client.post("/api/deals", json=dict(held_on="2026-06-19", client="A社", fee=400000))
    res = client.get("/api/export/excel", params={"fiscal_year": 2026})
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws.max_row >= 2
